import os
import io
import msoffcrypto
from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.utils import get_column_letter


class SpreadsheetManager:

    def __init__(self, file_path, is_protected=False, password=None):
        self.file_path = file_path
        self.is_protected = is_protected
        self._password = password
        self._wb = None
        if self.is_protected and not self._password:
            raise Exception("need password if a file is protected")

    def create_spreadsheet(self):
        if os.path.exists(self.file_path):
            raise FileExistsError(f"File '{self.file_path}' already exists.")
        wb = Workbook()
        wb.save(self.file_path)
        print(f"Spreadsheet created at '{self.file_path}'.")

    def read_sheet(self, sheet_name):
        wb = self._load_workbook()
        sheet = wb[sheet_name] if sheet_name else wb.active
        data = []
        # headers = [cell.value for cell in sheet[1]]
        for row in sheet.iter_rows(min_row=1, values_only=True):
            data.append(row)
        return data

    def add_sheet(self, sheet_name):
        wb = self._load_workbook()
        if sheet_name in wb.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' already exists.")
        wb.create_sheet(sheet_name)
        wb.save(self.file_path)
        print(f"Sheet '{sheet_name}' added to '{self.file_path}'.")

    def update_sheet(self, sheet_name, data):
        wb = self._load_workbook()
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' does not exist.")
        ws = wb[sheet_name]
        for row in data:
            ws.append(row)
        wb.save(self.file_path)
        print(f"Sheet '{sheet_name}' updated in '{self.file_path}'.")

    def rename_sheet(self, old_name, new_name):
        wb = self._load_workbook()
        if old_name not in wb.sheetnames:
            raise ValueError(f"Sheet '{old_name}' does not exist.")
        if new_name in wb.sheetnames:
            raise ValueError(f"Sheet '{new_name}' already exists.")
        ws = wb[old_name]
        ws.title = new_name
        wb.save(self.file_path)
        print(f"Sheet '{old_name}' renamed to '{new_name}'.")

    def list_sheets(self):
        wb = self._load_workbook()
        return wb.sheetnames

    def delete_sheet(self, sheet_name):
        wb = self._load_workbook()
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' does not exist.")
        del wb[sheet_name]
        wb.save(self.file_path)
        print(f"Sheet '{sheet_name}' deleted from '{self.file_path}'.")

    def delete_spreadsheet(self):
        if os.path.exists(self.file_path):
            os.remove(self.file_path)
            print(f"Spreadsheet '{self.file_path}' deleted.")
        else:
            print(f"Spreadsheet '{self.file_path}' does not exist.")

    def create_graph(
        self,
        sheet_name,
        chart_type,
        data_range,
        categories_range,
        chart_title,
        width=30,
        height=15,
    ):
        wb = self._load_workbook()
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' does not exist.")
        ws = wb[sheet_name]

        if chart_type == "bar":
            chart = BarChart()
        elif chart_type == "line":
            chart = LineChart()
        elif chart_type == "pie":
            chart = PieChart()
        else:
            raise ValueError(f"Chart type '{chart_type}' is not supported.")

        data = Reference(ws, range_string=f"{sheet_name}!{data_range}")
        categories = Reference(ws, range_string=f"{sheet_name}!{categories_range}")

        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)

        chart.title = chart_title
        chart.width = width
        chart.height = height

        chart_location = f"{get_column_letter(len(list(ws.columns)) + 1)}1"
        ws.add_chart(chart, chart_location)

        wb.save(self.file_path)
        print(f"Chart '{chart_title}' created in '{sheet_name}'.")

    def _load_workbook(self):
        if not os.path.exists(self.file_path):
            raise FileNotFoundError(
                f"File '{self.file_path}' does not exist. Create it first."
            )
        if self._wb:
            return self._wb
        if self.is_protected:
            try:
                with open(self.file_path, "rb") as file:
                    decrypt_file = msoffcrypto.OfficeFile(file)
                    decrypt_file.load_key(password=self._password)
                    decrypted = io.BytesIO()
                    decrypt_file.decrypt(decrypted)
                    decrypted.seek(0)
                    self._wb = load_workbook(decrypted)
                    return self._wb
            except Exception as e:
                raise Exception(f"Error loading or decrypting the file: {e}")
        else:
            self._wb = load_workbook(self.file_path)
            return self._wb
